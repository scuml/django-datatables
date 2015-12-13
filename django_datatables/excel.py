import openpyxl
from openpyxl.styles import Style, Font, Color
from openpyxl.writer.dump_worksheet import ExcelDumpWriter

from collections import defaultdict, OrderedDict
from HTMLParser import HTMLParser
from io import BytesIO
from zipfile import ZipFile, ZIP_DEFLATED

from django.http import HttpResponse


class MLStripper(HTMLParser):
    def __init__(self):
        self.reset()
        self.fed = []

    def handle_data(self, d):
        self.fed.append(d)

    def get_data(self):
        return ''.join(self.fed)

def strip_tags(html):
    s = MLStripper()
    s.feed(html)
    return s.get_data()

class ExcelWriter(object):
    """
    Data structure of sheets with headers and data
    that allows for consistent export of data to excel.
    """

    def __init__(self):
        self.sheets = defaultdict(lambda: defaultdict(
            headers=list(),
            data=list(),
        ))
        self.workbook = openpyxl.Workbook(optimized_write=True)

    def add_headers(self, sheet_name, headers_list):
        """
        Add headers to receipt.
        :param sheet_name: name of sheet
        :param headers_list: list of column headers
        """
        self.sheets[sheet_name]["headers"].extend(headers_list)

    def set_data(self, all_data):
        """ Adds all data as {sheet_name:list(data)}"""
        for sheet_name, data in all_data.items():
            self.sheets[sheet_name]['data'] = data

    def set_sheet_data(self, sheet_name, data):
        """ Set a specific sheet to a list of data """
        self.sheets[sheet_name]['data'] = data

    def add_row(self, sheet_name, data_dict):
        """
        Append a row of data to a sheet
        :param sheet_name: name of sheet
        :param data_dict:
            dict where key, value --> col header name, row val to save
        """
        self.sheets[sheet_name]['data'].append(data_dict)

    def _get_row_style(self, row_dict):
        """ Override this function with logic
        that returns row style based on rec's contents
        :param row_dict: dict where key, value --> col header, row's value
        :return: None or an openpyxl.styles.Style
        """
        return None

    def _create_sheets(self):
        """
        Creates sheets in workbook using current sheets data
        """
        header_style = Style(Font(bold=True))

        for sheet_name, data in self.sheets.items():
            sheet = self.workbook.create_sheet(title=sheet_name)

            header_row = list()
            for header in data['headers']:
                header_row.append(dict(
                    value=header,
                    style=header_style,
                ))
            sheet.append(header_row)
            row_num = 1

            for rec in iter(data["data"]):
                if rec is None:
                    continue

                outrow = list()
                for index, key in enumerate(data['headers']):
                    try:
                        value = rec[key]
                    except KeyError:
                        value = ''
                    if isinstance(value, list):
                        value = '\r\n'.join(value)
                    if value is None:
                        value = ''
                    if not isinstance(value, unicode):
                        value = unicode(value).decode('utf-8')

                    value = strip_tags(value)

                    cell = dict(
                        value=value,
                        style=self._get_row_style(rec),
                    )

                    outrow.append(cell)

                sheet.append(outrow)
                row_num += 1

    def _save_virtual_workbook(self):
        """
        Return an in-memory workbook, suitable for a Django response.

        NOTE: Fixes openpyxl.writer.excel.save_virtual_workbook() for
        optimized workbooks. Package function uses ExcelWriter instead of
        ExcelDumpWriter with no way of specifying it.
        Had to copy the function and change which Writer it used.
        """
        writer = ExcelDumpWriter(self.workbook)
        temp_buffer = BytesIO()
        try:
            archive = ZipFile(temp_buffer, 'w', ZIP_DEFLATED)
            writer.write_data(archive)
        finally:
            archive.close()
        virtual_workbook = temp_buffer.getvalue()
        temp_buffer.close()
        return virtual_workbook

    def download(self, filename):
        """
        Return an HttpResponse downloading the workbook
        """
        self._create_sheets()

        response = HttpResponse(
            self._save_virtual_workbook(),
            content_type='application/vnd.ms-excel'
        )
        response['Content-Disposition'] = \
            'attachment; filename="{0}"'.format(filename)

        return response

    def save_to_file(self, filename):
        """"
        Save workbook to a file
        """
        self._create_sheets()
        self.workbook.save(filename)
